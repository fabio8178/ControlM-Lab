from openpyxl import load_workbook
from pathlib import Path
from job import Job
from folder import Folder
from config_reader import carregar_config


def ler_jobs():
    config = carregar_config()
    arquivo = Path(config["input_file"])
    
    if not arquivo.exists():
        print("Arquivo Excel não encontrado!")
        return []

    workbook = load_workbook(arquivo)
    planilha = workbook.active

    folders = []

    for linha in planilha.iter_rows(min_row=2, values_only=True):

        folder_encontrada = None

        # Procura se a Folder já existe
        for folder in folders:

            if folder.nome == linha[0]:
                folder_encontrada = folder
                break

        # Se não encontrou, cria uma nova
        if folder_encontrada is None:

            folder_encontrada = Folder(
                linha[0],  # Nome da Folder
                linha[2]   # Application
            )

            folders.append(folder_encontrada)

        # Cria o Job
        job = Job(
            linha[1],
            linha[2],
            linha[3],
            linha[4],
            linha[5],
            linha[6],
            linha[7],
            linha[8]
        )

        # Adiciona o Job na Folder encontrada
        folder_encontrada.jobs.append(job)

    return folders